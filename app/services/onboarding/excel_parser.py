"""Excel parsing for the onboarding agent.

Handles .xlsx (and .xls when openpyxl reads them, otherwise falls back) with
real-world quirks:

* Header row not necessarily in row 1 (e.g. summary rows above)
* Merged cells in headers — value propagated to all merged positions
* Empty leading/trailing columns
* Multiple sheets — picks the most "data-like" sheet by default

The output is JSON-serialisable so it can be persisted in
``OnboardingSession.state_json``.
"""

from __future__ import annotations

import datetime
import io
from dataclasses import dataclass
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class ParsedSheet:
    """Result of parsing a single Excel sheet."""

    sheet_name: str
    headers: list[str]
    sample_rows: list[
        list[Any]
    ]  # JSON-safe cell values (str / int / float / None / iso date)
    total_data_rows: int  # excluding header
    header_row_index: int  # 1-based row number in the worksheet


@dataclass(frozen=True)
class ParsedWorkbook:
    """All data-like sheets parsed from a single Excel workbook."""

    sheets: list[ParsedSheet]


# Maximum number of sample rows to keep — these are the rows we will send to
# the LLM (after anonymisation) and also the rows we show in the preview.
MAX_SAMPLE_ROWS = 5


def parse_workbook(content: bytes) -> ParsedWorkbook:
    """Parse every sheet that looks like a data table.

    A workbook frequently contains one sheet per (plot, campaign) plus a
    few auxiliary sheets (``RESUMEN``, ``GASTOS``, ``Hoja3``). We only
    return sheets where the parser was able to find a header row and at
    least one data row with 3+ columns. The order of :pyattr:`sheets`
    matches the order of sheets in the workbook.
    """
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=False)
    except Exception as exc:  # pragma: no cover - openpyxl errors vary
        raise ValueError(f"No se pudo leer el fichero Excel: {exc}") from exc
    if not wb.sheetnames:
        raise ValueError("El fichero Excel no tiene hojas.")

    sheets: list[ParsedSheet] = []
    for sheet_name in wb.sheetnames:
        try:
            ps = parse_excel(content, sheet_name=sheet_name)
        except ValueError:
            continue
        if ps.total_data_rows < 1 or len(ps.headers) < 3:
            continue
        sheets.append(ps)
    if not sheets:
        raise ValueError(
            "Ninguna hoja del fichero tiene una estructura tabular reconocible."
        )
    return ParsedWorkbook(sheets=sheets)


def parse_excel(content: bytes, sheet_name: str | None = None) -> ParsedSheet:
    """Parse an Excel file and return headers + sample rows.

    Args:
        content: Raw bytes of an .xlsx file.
        sheet_name: Optional sheet to pick. If ``None`` the heuristic in
            :func:`_pick_sheet` is used.

    Returns:
        A :class:`ParsedSheet` with JSON-safe values.

    Raises:
        ValueError: if the file cannot be parsed as an Excel workbook or no
            usable header row is found.
    """
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=False)
    except Exception as exc:  # pragma: no cover - openpyxl errors vary
        raise ValueError(f"No se pudo leer el fichero Excel: {exc}") from exc

    if not wb.sheetnames:
        raise ValueError("El fichero Excel no tiene hojas.")

    target = sheet_name or _pick_sheet(wb)
    if target not in wb.sheetnames:
        raise ValueError(f"La hoja '{target}' no existe en el fichero.")

    ws = wb[target]
    _unmerge_with_propagation(ws)

    header_row_idx, headers = _detect_header_row(ws)
    if not headers:
        raise ValueError(
            "No se ha podido detectar una fila de cabecera con suficientes "
            "columnas en la hoja '{sheet}'.".format(sheet=target)
        )
    headers = _dedupe_headers(headers)

    sample_rows: list[list[Any]] = []
    total = 0
    n_cols = len(headers)
    consecutive_empty = 0
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        # Trim/pad row to header length
        values = list(row[:n_cols]) + [None] * max(0, n_cols - len(row))
        if not any(_is_present(v) for v in values):
            consecutive_empty += 1
            # Stop after 5 consecutive empty rows — data block clearly ended.
            if consecutive_empty >= 5:
                break
            continue
        consecutive_empty = 0
        if _is_footer_row(values):
            # Summary / totals / gastos section below the data — stop here.
            break
        total += 1
        if len(sample_rows) < MAX_SAMPLE_ROWS:
            sample_rows.append([_to_json_safe(v) for v in values])

    return ParsedSheet(
        sheet_name=target,
        headers=headers,
        sample_rows=sample_rows,
        total_data_rows=total,
        header_row_index=header_row_idx,
    )


# --- internal helpers ------------------------------------------------------


def _pick_sheet(wb) -> str:
    """Pick the sheet with the most non-empty cells (rough but effective)."""
    best_name = wb.sheetnames[0]
    best_score = -1
    for name in wb.sheetnames:
        ws = wb[name]
        score = 0
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if _is_present(v):
                    score += 1
                    if score > 200:
                        break
            if score > 200:
                break
        if score > best_score:
            best_score = score
            best_name = name
    return best_name


def _unmerge_with_propagation(ws: Worksheet) -> None:
    """Unmerge all merged ranges, propagating the anchor value to every cell.

    openpyxl by default keeps the value only in the top-left anchor cell when
    cells are merged, so iter_rows returns ``None`` for the rest. We unmerge
    and copy the value to all positions so downstream parsing sees the same
    value the user saw in Excel.
    """
    ranges = list(ws.merged_cells.ranges)
    for rng in ranges:
        min_col, min_row, max_col, max_row = rng.bounds
        anchor_value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(rng))
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if r == min_row and c == min_col:
                    continue
                ws.cell(row=r, column=c).value = anchor_value


def _detect_header_row(ws: Worksheet, max_scan_rows: int = 20) -> tuple[int, list[str]]:
    """Find the row that best looks like a header.

    Strategy: scan the first ``max_scan_rows`` rows and score each one as a
    header candidate. A header row typically has:

    * Several non-empty cells (>= 2).
    * Most cells are short text labels (not numbers/dates).
    * The rows immediately below contain mostly numeric/date cells (data).

    We also reject rows whose non-empty cells all share the same value
    (typical of a banner that spans multiple columns after we propagate the
    merged value).
    """
    rows: list[tuple[int, list[Any]]] = []
    for ridx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1
    ):
        rows.append((ridx, list(row)))

    candidates: list[tuple[int, int, int, list[Any]]] = []
    for i, (ridx, row) in enumerate(rows):
        non_empty = [v for v in row if _is_present(v)]
        if len(non_empty) < 2:
            continue
        # Skip rows that are just a banner value repeated across columns.
        text_values = [v for v in non_empty if isinstance(v, str)]
        if (
            text_values
            and len(text_values) == len(non_empty)
            and len({v.strip() for v in text_values}) == 1
        ):
            continue
        # Count short text labels vs numeric/date data cells.
        text_count = sum(
            1
            for v in non_empty
            if isinstance(v, str) and 1 <= len(v.strip()) <= 60
        )
        numeric_count = sum(
            1
            for v in non_empty
            if isinstance(v, (int, float, datetime.datetime, datetime.date))
            and not isinstance(v, bool)
        )
        if text_count < 2 or numeric_count > text_count:
            continue
        # Look at the next few rows: data rows boost the score.
        data_score = 0
        for j in range(i + 1, min(i + 5, len(rows))):
            _, nxt = rows[j]
            for v in nxt:
                if isinstance(v, (int, float, datetime.datetime, datetime.date)) and not isinstance(
                    v, bool
                ):
                    data_score += 1
        score = text_count * 3 + data_score
        candidates.append((score, -ridx, ridx, row))

    if not candidates:
        return 0, []
    candidates.sort(reverse=True)
    _, _, ridx, row = candidates[0]
    trimmed = list(row)
    while trimmed and not _is_present(trimmed[-1]):
        trimmed.pop()
    headers = [_normalize_header(v, i) for i, v in enumerate(trimmed)]
    return ridx, headers


# Keywords that mark the start of a summary / totals / secondary section in
# Spanish/English. When the first non-empty cell of a row matches one of
# these (or starts with it) we stop reading the data block.
_FOOTER_KEYWORDS: frozenset[str] = frozenset(
    {
        "TOTAL",
        "TOTALES",
        "SUMA",
        "SUBTOTAL",
        "SALDO",
        "SALDOS",
        "GASTOS",
        "INGRESOS",
        "RESUMEN",
        "PROMEDIO",
        "MEDIA",
        "BALANCE",
        "TOTAL GENERAL",
    }
)


def _is_footer_row(values: list[Any]) -> bool:
    """Return True if the row looks like the start of a summary block."""
    first_non_empty = next((v for v in values if _is_present(v)), None)
    if not isinstance(first_non_empty, str):
        return False
    s = first_non_empty.strip().upper().rstrip(":=").strip()
    if not s:
        return False
    if s in _FOOTER_KEYWORDS:
        return True
    # Accept "TOTAL ...", "SALDO GENERAL", "TOTAL 2025", etc.
    first_token = s.split()[0]
    return first_token in _FOOTER_KEYWORDS


def _dedupe_headers(headers: list[str]) -> list[str]:
    """Ensure header names are unique by suffixing duplicates with ``(2)``,
    ``(3)``, etc. Needed because the resolved mapping is keyed by header
    name in the UI form.
    """
    seen: dict[str, int] = {}
    out: list[str] = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            out.append(f"{h} ({seen[h]})")
        else:
            seen[h] = 1
            out.append(h)
    return out


def _normalize_header(value: Any, idx: int) -> str:
    if value is None or (isinstance(value, str) and not value.strip()):
        return f"col_{idx + 1}"
    if isinstance(value, str):
        return value.strip()
    return str(value)


def _is_present(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str) and not value.strip():
        return False
    return True


def _to_json_safe(value: Any) -> Any:
    """Convert a cell value to something JSON-serialisable."""
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.isoformat()
    # Fallback: stringify (openpyxl may yield Decimal, timedelta, etc.)
    return str(value)
