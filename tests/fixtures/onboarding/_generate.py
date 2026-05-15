"""Generates synthetic Excel fixtures used by the onboarding tests.

Run from the repository root::

    uv run python tests/fixtures/onboarding/_generate.py

Three fixtures are produced:

* ``parcelas_limpio.xlsx``               — straight tabular data in row 1.
* ``gastos_cabeceras_irregulares.xlsx``  — title rows above the header.
* ``ingresos_celdas_fusionadas.xlsx``    — merged header cells.
"""

from __future__ import annotations

import datetime
from pathlib import Path

from openpyxl import Workbook

OUT_DIR = Path(__file__).resolve().parent


def _write_parcelas_clean() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Parcelas"
    ws.append(
        [
            "Nombre",
            "Fecha plantación",
            "Polígono",
            "Parcela",
            "Hidrante",
            "Sector",
            "Nº plantas",
            "Superficie (ha)",
        ]
    )
    rows = [
        (
            "Via Minera Javi",
            datetime.date(2019, 11, 1),
            "12",
            "265",
            "H340",
            "3",
            134,
            0.5555,
        ),
        (
            "Santa Cruz Mar",
            datetime.date(2013, 9, 1),
            "29",
            "134",
            "H331",
            "3",
            122,
            0.5112,
        ),
        ("La Hoya", datetime.date(2018, 4, 15), "8", "112", "H210", "1", 95, 0.4321),
    ]
    for r in rows:
        ws.append(r)
    wb.save(OUT_DIR / "parcelas_limpio.xlsx")


def _write_gastos_irregular() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos 2025"
    ws["A1"] = "INFORME ANUAL DE GASTOS"
    ws["A2"] = "Campaña 2025/26"
    ws["A3"] = ""  # blank
    ws.append(
        ["Fecha", "Descripción", "Proveedor", "Finca", "Importe (€)", "Categoría"]
    )
    rows = [
        (
            datetime.date(2025, 11, 14),
            "Pienso perros",
            "Cooperativa",
            "",
            21.0,
            "Perros",
        ),
        (
            datetime.date(2025, 12, 18),
            "Adeudo riego",
            "Comunidad regantes",
            "Santa Cruz Mar",
            105.0,
            "Riego",
        ),
        (
            datetime.date(2026, 1, 10),
            "Reparación motobomba",
            "Talleres Pérez",
            "La Hoya",
            230.50,
            "Mantenimiento",
        ),
        (datetime.date(2026, 2, 3), "Combustible", "Repsol", "", 64.20, "Suministros"),
    ]
    for r in rows:
        ws.append(r)
    wb.save(OUT_DIR / "gastos_cabeceras_irregulares.xlsx")


def _write_ingresos_merged() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"
    # Merged title spanning header columns
    ws["A1"] = "Ventas de trufa — campaña 2025/26"
    ws.merge_cells("A1:E1")
    # Header in row 3
    ws.append([])  # row 2
    ws.append(["Fecha", "Bancal", "Kg", "Calidad", "€/Kg"])
    rows = [
        (datetime.date(2025, 11, 13), "Santa Cruz Mar", 0.900, "", 45.0),
        (datetime.date(2025, 12, 3), "Santa Cruz Mar", 0.200, "A", 450.0),
        (datetime.date(2026, 1, 14), "Via Minera Javi", 1.250, "B", 320.0),
        (datetime.date(2026, 1, 28), "La Hoya", 0.450, "extra", 480.0),
    ]
    for r in rows:
        ws.append(r)
    wb.save(OUT_DIR / "ingresos_celdas_fusionadas.xlsx")


def _write_ingresos_real_world() -> None:
    """Inspired by a real truffle grower's Excel: two side-by-side tables
    (current + historical campaign), title row at the top, summary rows
    ("TOTAL", "GASTOS", "SUMA", "SALDO") below the data block, and
    duplicate column names between the two tables.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Ingresos 25-26"
    # Row 1: two campaign banners side by side + a stray date.
    ws.cell(row=1, column=1, value="CAMPAÑA 2025-2026")
    ws.cell(row=1, column=10, value="CAMPAÑA 2019-2020")
    ws.cell(row=1, column=12, value=datetime.datetime(2020, 5, 1))
    # Row 2: empty separator.
    # Row 3: real header for both tables.
    header = [
        "Fecha",
        "Tip",
        "Kg",
        "Precio",
        "Factura",
        "Otros Kg",
        "Otros Eur",
        "",
        "",
        "Fecha",
        "Tip",
        "Kg",
        "Precio",
        "Factura",
    ]
    for c, val in enumerate(header, start=1):
        ws.cell(row=3, column=c, value=val)
    # Rows 4..7: data.
    data = [
        (
            datetime.date(2025, 11, 8),
            "B",
            3.7,
            30,
            111,
            None,
            None,
            None,
            None,
            datetime.date(2019, 11, 9),
            "B",
            1.45,
            130,
            188.5,
        ),
        (
            datetime.date(2025, 11, 15),
            "B",
            3.4,
            50,
            170,
            None,
            None,
            None,
            None,
            datetime.date(2019, 11, 16),
            "B",
            4.25,
            200,
            850,
        ),
        (
            datetime.date(2025, 11, 22),
            "B",
            2.8,
            120,
            336,
            2.5,
            245,
            None,
            None,
            datetime.date(2019, 11, 23),
            "B",
            4.15,
            250,
            1037.5,
        ),
        (
            datetime.date(2025, 11, 29),
            "M",
            0.75,
            30,
            22.5,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
        ),
    ]
    for r in data:
        ws.append(list(r))
    # Row 8: TOTAL summary — parser must stop here.
    ws.append(["TOTAL", None, 10.65, None, 639.5])
    # Row 9: empty.
    ws.append([])
    # Row 10: GASTOS section header.
    ws.append(["GASTOS"])
    ws.append(["Labrar Chogo", None, None, None, 130])
    ws.append(["SUMA", None, None, None, 130])
    ws.append(["SALDO=", None, None, None, 509.5])
    wb.save(OUT_DIR / "ingresos_real_world.xlsx")


def _write_ingresos_multi_sheet() -> None:
    """Workbook with one Ingresos sheet per (plot, campaign) plus noise.

    Mirrors the structure of the real-world fixture in the prompt: the
    workbook contains several ``Ingresos PARCELA YY-YY`` sheets that share
    headers but cover different campaigns, plus a ``RESUMEN`` summary sheet
    and an empty ``Hoja3`` that should be filtered out.
    """
    wb = Workbook()
    # Drop the default sheet — we'll create our own in order.
    default = wb.active
    wb.remove(default)

    def _make_ingresos_sheet(title: str, rows: list[tuple]) -> None:
        ws = wb.create_sheet(title=title)
        ws.append(["Fecha", "Cliente", "Kg", "€/kg", "Importe"])
        for r in rows:
            ws.append(list(r))

    _make_ingresos_sheet(
        "Ingresos CERRELLAR 25-26",
        [
            (datetime.date(2025, 12, 1), "Cliente A", 1.2, 800, 960),
            (datetime.date(2025, 12, 15), "Cliente B", 0.8, 750, 600),
        ],
    )
    _make_ingresos_sheet(
        "Ingresos CERRELLAR 24-25",
        [
            (datetime.date(2024, 11, 20), "Cliente A", 0.5, 700, 350),
            (datetime.date(2025, 1, 10), "Cliente C", 1.5, 720, 1080),
            (datetime.date(2025, 2, 3), "Cliente B", 0.3, 700, 210),
        ],
    )

    # Summary sheet: only one column, must be skipped by parse_workbook.
    resumen = wb.create_sheet(title="RESUMEN")
    resumen.append(["Total temporada"])
    resumen.append([2200])

    # Empty noise sheet.
    wb.create_sheet(title="Hoja3")

    wb.save(OUT_DIR / "ingresos_multi_sheet.xlsx")


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    _write_parcelas_clean()
    _write_gastos_irregular()
    _write_ingresos_merged()
    _write_ingresos_real_world()
    _write_ingresos_multi_sheet()
    print(f"Fixtures generadas en {OUT_DIR}")


if __name__ == "__main__":
    main()
