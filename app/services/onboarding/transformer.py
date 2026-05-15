"""Transform raw Excel data + resolved mapping into Trufiq CSV format.

Given:
* the raw Excel ``content`` bytes (re-parsed to read all rows, not just the
  sample),
* the ``entity_type`` chosen by the user,
* the ``resolved_mapping`` (a list of ``{source_column, target_field}`` dicts),

produce a UTF-8 string with one CSV line per valid input row, semicolon
delimited, no header, in the exact column order expected by
``app.services.import_service``.

Best-effort coercion:
* dates: ``dateutil.parser.parse(dayfirst=True)`` → ``DD/MM/YYYY``
* numbers: accept either EU (``1.250,50``) or US (``1,250.50``) format → EU
* strings: ``str(...).strip()``

Rows whose required field cannot be coerced are skipped and reported as
``ValidationError`` entries.
"""

from __future__ import annotations

import csv
import io
from datetime import date, datetime
from typing import Any

from dateutil import parser as date_parser
from openpyxl import load_workbook

from app.services.onboarding.entity_schemas import EntitySchema, get_schema
from app.services.onboarding.excel_parser import _is_footer_row
from app.services.onboarding.state import ValidationError


def _coerce_date(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    parsed = date_parser.parse(str(value), dayfirst=True)
    return parsed.strftime("%d/%m/%Y")


def _coerce_number(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, (int, float)):
        # Render as EU number with comma decimal separator
        if isinstance(value, int):
            return str(value)
        return f"{value:.4f}".rstrip("0").rstrip(".").replace(".", ",")
    text = str(value).strip()
    # Detect European format (comma as decimal): "1.250,50"
    has_comma = "," in text
    has_dot = "." in text
    if has_comma and has_dot:
        # Assume EU: dots = thousands, comma = decimal
        normalized = text.replace(".", "").replace(",", ".")
    elif has_comma and not has_dot:
        # EU short form: 1250,50
        normalized = text.replace(",", ".")
    else:
        normalized = text
    f = float(normalized)
    if f.is_integer():
        return str(int(f))
    return f"{f:.4f}".rstrip("0").rstrip(".").replace(".", ",")


def _coerce_integer(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value))
    return str(int(float(str(value).replace(",", "."))))


def _coerce_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


_COERCERS = {
    "date": _coerce_date,
    "number": _coerce_number,
    "integer": _coerce_integer,
    "text": _coerce_text,
    "enum": _coerce_text,
    "boolean": _coerce_text,
}


def _read_all_rows(
    content: bytes, sheet_name: str | None, header_row_index: int
) -> list[list[Any]]:
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    rows: list[list[Any]] = []
    consecutive_empty = 0
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx <= header_row_index:
            continue
        values = list(row)
        if not any(cell is not None and cell != "" for cell in values):
            consecutive_empty += 1
            if consecutive_empty >= 5:
                break
            continue
        consecutive_empty = 0
        if _is_footer_row(values):
            break
        rows.append(values)
    return rows


def _column_index_map(
    headers: list[str], resolved_mapping: list[dict]
) -> dict[str, int]:
    """Map ``target_field`` → column index in the source row."""
    out: dict[str, int] = {}
    for entry in resolved_mapping:
        target = entry.get("target_field")
        source = entry.get("source_column")
        if not target or target == "IGNORE" or not source:
            continue
        if source in headers:
            out[target] = headers.index(source)
    return out


def transform_to_csv(
    *,
    content: bytes,
    sheet_name: str | None,
    headers: list[str],
    header_row_index: int,
    entity_type: str,
    resolved_mapping: list[dict],
    constants: dict[str, Any] | None = None,
) -> tuple[str, list[ValidationError]]:
    """Build a Trufiq-compatible CSV from the raw Excel bytes.

    ``constants`` provides default values for target fields that are not
    backed by a source column (e.g. the parcel name inferred from the
    sheet title). Per-row values from the mapping always take precedence
    when they are present.

    Returns a tuple ``(csv_text, errors)``. ``csv_text`` only contains the
    rows that could be coerced cleanly; problematic rows are reported via
    ``errors`` and skipped from the CSV.
    """
    schema: EntitySchema | None = get_schema(entity_type)
    if schema is None:
        raise ValueError(f"Esquema desconocido: {entity_type!r}")

    target_to_col = _column_index_map(headers, resolved_mapping)
    raw_rows = _read_all_rows(content, sheet_name, header_row_index)
    constants = constants or {}

    field_order = [f.id for f in schema.fields]
    required_ids = set(schema.required_field_ids())
    errors: list[ValidationError] = []

    out_buf = io.StringIO()
    writer = csv.writer(out_buf, delimiter=";", lineterminator="\n")

    for r_idx, raw in enumerate(raw_rows, start=1):
        out_row: list[str] = []
        row_failed = False
        for field_id in field_order:
            spec = schema.get(field_id)
            assert spec is not None  # by construction
            col_idx = target_to_col.get(field_id)
            if col_idx is None or col_idx >= len(raw):
                value = constants.get(field_id)
            else:
                value = raw[col_idx]
                if (value is None or value == "") and field_id in constants:
                    value = constants[field_id]

            if (value is None or value == "") and field_id in required_ids:
                errors.append(
                    {
                        "row_index": r_idx,
                        "column": field_id,
                        "message": f"Campo obligatorio '{spec.label_es}' vacío",
                    }
                )
                row_failed = True
                break

            try:
                coerced = _COERCERS.get(spec.type, _coerce_text)(value)
            except (ValueError, TypeError) as exc:
                errors.append(
                    {
                        "row_index": r_idx,
                        "column": field_id,
                        "message": f"No se pudo convertir '{value}' a {spec.type}: {exc}",
                    }
                )
                row_failed = True
                break
            out_row.append(coerced)

        if not row_failed:
            writer.writerow(out_row)

    return out_buf.getvalue(), errors
